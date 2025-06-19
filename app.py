from flask import Flask, render_template, request, redirect, session, send_file
from flask import flash
from datetime import datetime, timedelta
from models import db
from models.user import User
from models.shift import Shift
from models.schedule import Schedule
from scheduler.logic import generate_schedule
from flask import render_template, request, send_file
from datetime import datetime, timedelta
from models import User, Shift, Schedule
import openpyxl
from io import BytesIO
from flask import render_template, request
from datetime import datetime, timedelta, date
from models import User, Schedule, Shift, db

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.secret_key = 'lichtruc2025'
db.init_app(app)

@app.before_request
def require_login():
    allowed_routes = ['login']
    if 'user_id' not in session and request.endpoint not in allowed_routes:
        return redirect('/login')

@app.context_processor
def inject_user():
    user = None
    if 'user_id' in session:
        user = User.query.get(session['user_id'])
    return dict(user=user)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        print(f">>> Thử đăng nhập: {username} / {password}")
        user = User.query.filter_by(username=username).first()
        if user and user.password == password:
            session['user_id'] = user.id
            session['role'] = user.role
            session['department'] = user.department
            return redirect('/')
        return "Sai tài khoản hoặc mật khẩu."
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

@app.route('/assign', methods=['GET', 'POST'])
def assign_schedule():
    user_role = session.get('role')
    user_dept = session.get('department')
    if user_role != 'admin':
        departments = [user_dept]
    else:
        departments = [d[0] for d in db.session.query(User.department).filter(User.department != None).distinct().all()]

    selected_department = request.args.get('department') if request.method == 'GET' else request.form.get('department')
    users = User.query.filter_by(department=selected_department).all() if selected_department else []
    shifts = Shift.query.all()

    if request.method == 'POST':
        start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d').date()
        end_date = datetime.strptime(request.form['end_date'], '%Y-%m-%d').date()
        for checkbox in request.form.getlist('schedule'):
            user_id, shift_id = checkbox.split('-')
            current = start_date
            while current <= end_date:
                s = Schedule(user_id=int(user_id), shift_id=int(shift_id), work_date=current)
                db.session.add(s)
                current += timedelta(days=1)
        db.session.commit()
        return redirect(f'/assign?department={selected_department}')

    return render_template('assign.html', departments=departments, selected_department=selected_department, users=users, shifts=shifts)

@app.route('/auto-assign')
def auto_assign_page():
    selected_department = request.args.get('department')

    departments = db.session.query(User.department).distinct().all()
    departments = [d[0] for d in departments if d[0]]

    users = User.query.filter_by(department=selected_department).all() if selected_department else []
    shifts = Shift.query.all()

    return render_template('auto_assign.html',
                           departments=departments,
                           selected_department=selected_department,
                           users=users,
                           shifts=shifts)

@app.route('/schedule', methods=['GET', 'POST'])
def view_schedule():
    selected_department = request.args.get('department')
    if session.get('role') in ['manager', 'user']:
        selected_department = session.get('department')

    departments = [d[0] for d in db.session.query(User.department).filter(User.department != None).distinct().all()]
    query = Schedule.query.join(User).join(Shift)

    if selected_department:
        query = query.filter(User.department == selected_department)

    # Lọc theo khoảng ngày nếu có
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        query = query.filter(Schedule.work_date.between(start_date, end_date))
    else:
        # nếu không chọn thì lấy 7 ngày tới mặc định
        start_date = datetime.today().date()
        end_date = start_date + timedelta(days=6)
        query = query.filter(Schedule.work_date.between(start_date, end_date))

    schedules = query.order_by(Schedule.work_date).all()

    # Tập hợp ngày trực
    date_range = sorted({s.work_date for s in schedules})

    # Pivot dữ liệu thành người → ngày → ca trực
    schedule_data = {}
    for s in schedules:
        u = s.user
        if u.id not in schedule_data:
            schedule_data[u.id] = {
                'id': u.id,
                'name': u.name,
                'position': u.position,
                'department': u.department,
                'shifts': {},
                'shifts_full': {}  # ✅ để template hiển thị và xóa từng ca
            }
        # Gán dữ liệu ca trực dạng đầy đủ
        schedule_data[u.id]['shifts'][s.work_date] = s.shift.name
        schedule_data[u.id]['shifts_full'][s.work_date] = {
            'shift_id': s.shift.id,
            'shift_name': s.shift.name
        }

    return render_template('schedule.html',
                           departments=departments,
                           selected_department=selected_department,
                           schedule_data=schedule_data,
                           date_range=date_range,
                           start_date=start_date,
                           end_date=end_date,
                           now=datetime.now())

@app.route('/schedule/edit/<int:user_id>', methods=['GET', 'POST'])
def edit_user_schedule(user_id):
    user = User.query.get_or_404(user_id)
    shifts = Shift.query.all()
    schedules = Schedule.query.filter_by(user_id=user_id).all()

    if request.method == 'POST':
        for s in schedules:
            new_shift_id = request.form.get(f'shift_{s.id}')
            if new_shift_id and int(new_shift_id) != s.shift_id:
                s.shift_id = int(new_shift_id)
        db.session.commit()
        return redirect('/schedule')

    return render_template('edit_schedule.html', user=user, shifts=shifts, schedules=schedules)

@app.route('/schedule/delete-one', methods=['POST'])
def delete_one_schedule():
    if session.get('role') not in ['admin', 'manager']:
        return "Bạn không có quyền xoá ca trực.", 403

    user_id = request.form.get('user_id')
    shift_id = request.form.get('shift_id')
    work_date = request.form.get('work_date')

    schedule = Schedule.query.filter_by(
        user_id=user_id,
        shift_id=shift_id,
        work_date=work_date
    ).first()

    if schedule:
        db.session.delete(schedule)
        db.session.commit()

    return redirect(request.referrer or '/schedule')

@app.route('/calendar')
def fullcalendar():
    selected_department = request.args.get('department')
    if session.get('role') in ['manager', 'user']:
        selected_department = session.get('department')

    departments = [d[0] for d in db.session.query(User.department).filter(User.department != None).distinct().all()]
    query = Schedule.query.join(User)

    if selected_department:
        query = query.filter(User.department == selected_department)

    schedules = query.order_by(Schedule.work_date).all()
    return render_template('fullcalendar.html',
                           schedules=schedules,
                           departments=departments,
                           selected_department=selected_department)

@app.route('/stats')
def stats():
    from sqlalchemy import func

    user_role = session.get('role')
    user_dept = session.get('department')

    query_users = User.query
    query_schedules = Schedule.query.join(User)

    if user_role in ['manager', 'user']:
        query_users = query_users.filter(User.department == user_dept)
        query_schedules = query_schedules.filter(User.department == user_dept)


    total_users = query_users.count()
    total_shifts = Shift.query.count()
    total_schedules = query_schedules.count()

    schedules_per_user = db.session.query(User.name, func.count(Schedule.id))\
        .join(Schedule)\
        .filter(User.department == user_dept if user_role != 'admin' else True)\
        .group_by(User.id).all()

    return render_template('stats.html',
                           total_users=total_users,
                           total_shifts=total_shifts,
                           total_schedules=total_schedules,
                           schedules_per_user=schedules_per_user)

@app.route('/report-by-department')
def report_by_department():
    user_role = session.get('role')
    user_dept = session.get('department')

    query = Schedule.query.join(User).join(Shift)
    if user_role in ['manager', 'user']:
        query = query.filter(User.department == user_dept)

    schedules = query.all()
    report = {}

    for s in schedules:
        dept = s.user.department
        if dept not in report:
            report[dept] = []
        report[dept].append(s)

    return render_template('report_by_department.html', report=report)

@app.route('/users-by-department')
def users_by_department():
    if session.get('role') in ['manager', 'user']:
        users = User.query.filter_by(department=session.get('department')).all()
        departments = [session.get('department')]
        selected_department = session.get('department')
    else:
        departments = [d[0] for d in db.session.query(User.department).filter(User.department != None).distinct().all()]
        selected_department = request.args.get('department')
        if selected_department:
            users = User.query.filter_by(department=selected_department).all()
        else:
            users = User.query.order_by(User.department).all()

    return render_template('users_by_department.html',
                           users=users,
                           departments=departments,
                           selected_department=selected_department)

@app.route('/export-by-department', methods=['GET', 'POST'])
def export_by_department():
    from sqlalchemy import distinct

    user_role = session.get('role')
    user_dept = session.get('department')

    # Lấy danh sách khoa
    departments = [d[0] for d in db.session.query(distinct(User.department)).filter(User.department != None).all()]
    selected_department = request.form.get('department') if request.method == 'POST' else user_dept

    if user_role != 'admin':
        selected_department = user_dept

    # Tạo file Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Họ tên', 'Ca trực', 'Ngày trực'])

    # Truy vấn lịch có chứa từ "trực"
    query = Schedule.query.join(User).join(Shift).filter(Shift.name.ilike('%trực%'))

    if selected_department:
        query = query.filter(User.department == selected_department)

    schedules = query.order_by(Schedule.work_date).all()

    for s in schedules:
        if s.user and s.shift:
            ws.append([s.user.name, s.shift.name, s.work_date.strftime('%Y-%m-%d')])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(stream, as_attachment=True, download_name="lichtruc_theo_khoa.xlsx")

@app.route('/generate_schedule', methods=['GET', 'POST'])
def generate_schedule_route():
    try:
        department = request.form.get('department')
        start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d').date()
        end_date = datetime.strptime(request.form['end_date'], '%Y-%m-%d').date()
        user_ids = request.form.getlist('user_ids')
        shift_ids = request.form.getlist('shift_ids')

        if not user_ids or not shift_ids:
            flash("Vui lòng chọn ít nhất 1 người và 1 ca trực.")
            return redirect(request.referrer)

        # Chuẩn bị danh sách ngày
        date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

        # Xếp lịch tua
        assignments = []
        index = 0
        for d in date_range:
            for sid in shift_ids:
                uid = user_ids[index % len(user_ids)]
                assignments.append(Schedule(user_id=uid, shift_id=sid, work_date=d))
                index += 1

        # Lưu vào DB
        db.session.add_all(assignments)
        db.session.commit()

        flash("✅ Đã tạo lịch tự động thành công.")
        return redirect(url_for('assign_schedule'))

    except Exception as e:
        db.session.rollback()
        flash(f"Lỗi tạo lịch: {str(e)}")
        return redirect(request.referrer)

@app.route('/export')
def export_excel():
    user_role = session.get('role')
    user_dept = session.get('department')
    wb.active = wb.active  # Đảm bảo chọn đúng sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Họ tên', 'Ca trực', 'Ngày trực'])
    ws.freeze_panes = "A2"  # ✅ Cố định hàng tiêu đề

    # Lấy lịch trực có chứa từ "trực"
    query = Schedule.query.join(User).join(Shift).filter(Shift.name.ilike('%trực%'))
    if user_role != 'admin':
        query = query.filter(User.department == user_dept)

    schedules = query.order_by(Schedule.work_date).all()
    for s in schedules:
        if s.user and s.shift:
            ws.append([s.user.name, s.shift.name, s.work_date.strftime('%Y-%m-%d')])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(stream, as_attachment=True, download_name="lichtruc.xlsx")

@app.route('/shifts')
def list_shifts():
    shifts = Shift.query.all()
    return render_template('shifts.html', shifts=shifts)

@app.route('/shifts/add', methods=['GET', 'POST'])
def add_shift():
    if request.method == 'POST':
        name = request.form['name']
        code = request.form['code']
        start_time = datetime.strptime(request.form['start_time'], '%H:%M').time()
        end_time = datetime.strptime(request.form['end_time'], '%H:%M').time()
        duration = float(request.form['duration'])
        shift = Shift(name=name, code=code, start_time=start_time, end_time=end_time, duration=duration)
        db.session.add(shift)
        db.session.commit()
        return redirect('/shifts')
    return render_template('add_shift.html')
    print(">>> CODE FIELD:", request.form.get('code'))

@app.route('/shifts/edit/<int:shift_id>', methods=['GET', 'POST'])
def edit_shift(shift_id):
    shift = Shift.query.get_or_404(shift_id)
    if request.method == 'POST':
        shift.name = request.form['name']
        code = request.form['code']  # chu y
        shift.start_time = datetime.strptime(request.form['start_time'], '%H:%M').time()
        shift.end_time = datetime.strptime(request.form['end_time'], '%H:%M').time()
        shift.duration = float(request.form['duration'])
        db.session.commit()
        return redirect('/shifts')
    return render_template('edit_shift.html', shift=shift)

@app.route('/users/edit/<int:user_id>', methods=['GET', 'POST'])
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    if request.method == 'POST':
        new_username = request.form['username']
        if User.query.filter(User.username == new_username, User.id != user_id).first():
            error_message = "User này đã có người dùng, bạn không thể cập nhật."
            return render_template('edit_user.html', user=user, error=error_message)

        user.name = request.form['name']
        user.username = new_username
        user.password = request.form['password']
        user.role = request.form['role']
        user.department = request.form['department']
        user.position = request.form['position']
        user.contract_type = request.form.get('contract_type')  # ✅ nếu có thêm trường này
        user.phone = request.form['phone']
        user.email = request.form['email']
        db.session.commit()
        return redirect('/users-by-department')

    return render_template('edit_user.html', user=user)


@app.route('/users/add', methods=['GET', 'POST'])
def add_user():
    error_message = None
    if request.method == 'POST':
        name = request.form['name']
        username = request.form['username']
        password = request.form['password']
        role = request.form['role']
        department = request.form['department']
        position = request.form['position']
        contract_type = request.form.get('contract_type')  # ✅ THÊM DÒNG NÀY
        phone = request.form.get('phone')
        email = request.form.get('email')

        # Kiểm tra trùng username
        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            error_message = "User này đã có người dùng, bạn không thể tạo được."
            return render_template('add_user.html', error=error_message)
        
        new_user = User(
            name=name,
            username=username,
            password=password,
            role=role,
            department=department,
            position=position,
            contract_type=contract_type,  # ✅ THÊM DÒNG NÀY
            phone=phone,
            email=email
        )
        db.session.add(new_user)
        db.session.commit()
        return redirect('/users-by-department')
    return render_template('add_user.html')

@app.route('/import-users', methods=['GET', 'POST'])
def import_users():
    if request.method == 'POST':
        file = request.files['file']
        if file.filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                row = row[:8]  # Lấy đúng 8 cột đầu tiên
                if len(row) < 6:
                    continue
                name, username, password, role, department, position = row[:6]
                email = row[6] if len(row) > 6 else None
                phone = row[7] if len(row) > 7 else None

                user = User(
                    name=name,
                    username=username,
                    password=password,
                    role=role,
                    department=department,
                    position=position,
                    email=email,
                    phone=phone
                )
                db.session.add(user)
            db.session.commit()
            return redirect('/users-by-department')
        else:
            return "Vui lòng chọn file .xlsx"
    return render_template('import_users.html')

@app.route('/roles', methods=['GET', 'POST'])
def manage_roles():
    if session.get('role') != 'admin':
        return "Bạn không có quyền truy cập trang này."

    users = User.query.order_by(User.department).all()

    if request.method == 'POST':
        for user in users:
            role = request.form.get(f'role_{user.id}')
            dept = request.form.get(f'department_{user.id}')
            if role and dept:
                user.role = role
                user.department = dept
        db.session.commit()
        return redirect('/roles')

    departments = [d[0] for d in db.session.query(User.department).distinct().all() if d[0]]
    roles = ['admin', 'manager', 'user']  # cập nhật quyền hệ thống
    positions = ['Bác sĩ', 'Điều dưỡng', 'Kỹ thuật viên']  # chức danh chuyên môn
    return render_template('manage_roles.html', users=users, departments=departments, roles=roles, positions=positions)

@app.route('/users/delete/<int:user_id>', methods=['POST', 'GET'])
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    db.session.delete(user)
    db.session.commit()
    return redirect('/users-by-department')

@app.route('/export-template', methods=['POST'])
def export_template():
    department = request.form.get('department')
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')

    print(">>> [EXPORT] Khoa:", department)
    print(">>> [EXPORT] Từ ngày:", start_date)
    print(">>> [EXPORT] Đến ngày:", end_date)

    query = Schedule.query.join(User).join(Shift)

    if department:
        query = query.filter(User.department == department)

    if start_date and end_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(Schedule.work_date.between(start, end))
        except ValueError:
            return "Ngày không hợp lệ.", 400
    else:
        return "Vui lòng chọn khoảng thời gian.", 400

    schedules = query.order_by(Schedule.work_date).all()
    if not schedules:
        return "Không có dữ liệu lịch trực.", 404

    # Tập hợp ngày
    date_range = sorted({s.work_date for s in schedules})

    # Pivot dữ liệu người dùng
    schedule_data = {}
    for s in schedules:
        u = s.user
        if u.id not in schedule_data:
            schedule_data[u.id] = {
                'name': u.name,
                'position': u.position,
                'department': u.department,
                'shifts': {}
            }
        schedule_data[u.id]['shifts'][s.work_date] = s.shift.name

    # --- Tạo Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lịch trực ngang"

    # --- Quốc hiệu, tiêu đề đầu trang ---
    ws.merge_cells('A1:G1')
    ws['A1'] = "BỆNH VIỆN NHI TỈNH GIA LAI"
    ws.merge_cells('H1:N1')
    ws['H1'] = "CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM"
    ws.merge_cells('H2:N2')
    ws['H2'] = "Độc lập - Tự do - Hạnh phúc"
    ws.merge_cells('A4:N4')
    ws['A4'] = f"BẢNG LỊCH TRỰC KHOA {department.upper() if department else ''}"
    ws.merge_cells('A5:N5')
    ws['A5'] = f"Lịch trực tuần ngày {start.strftime('%d/%m/%Y')} đến ngày {end.strftime('%d/%m/%Y')}"

    # --- Dòng tiêu đề bảng bắt đầu từ dòng 7 ---
    start_row = 7
    header = ['Họ tên', 'Chức danh', 'Khoa'] + [d.strftime('%d/%m') for d in date_range]
    ws.append(header)

    # Dữ liệu từng người
    for u in schedule_data.values():
        row = [u['name'], u['position'], u['department']]
        for d in date_range:
            row.append(u['shifts'].get(d, ''))  # Nếu không có ca → để trống
        ws.append(row)

    # --- Chân trang ---
    last_row = ws.max_row + 2
    ws[f'A{last_row}'] = "Nơi nhận:"
    ws[f'A{last_row+1}'] = "- Ban Giám đốc"
    ws[f'A{last_row+2}'] = "- Các khoa/phòng"
    ws[f'A{last_row+3}'] = "- Đăng website"
    ws[f'A{last_row+4}'] = "- Lưu: VP, KH-CNTT"

    ws.merge_cells(start_row=last_row, start_column=5, end_row=last_row, end_column=7)
    ws.cell(row=last_row, column=5).value = "Người lập bảng"
    ws.merge_cells(start_row=last_row, start_column=10, end_row=last_row, end_column=12)
    ws.cell(row=last_row, column=10).value = "GIÁM ĐỐC"

    ws.cell(row=last_row+1, column=5).value = "(Ký, ghi rõ họ tên)"
    ws.cell(row=last_row+1, column=10).value = "(Ký, ghi rõ họ tên)"

    # --- Xuất file ---
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return send_file(stream, as_attachment=True, download_name="lichtruc_dangngang.xlsx")

from sqlalchemy import or_

@app.route('/bang-cham-cong')
def bang_cham_cong():
    from datetime import datetime, timedelta, date

    today = datetime.today()
    start_date = request.args.get('start', today.replace(day=1).strftime('%Y-%m-%d'))
    end_date = request.args.get('end', today.strftime('%Y-%m-%d'))
    department = request.args.get('department', '')
    selected_contract = request.args.get('contract_type', '')
    print_filter = request.args.get('print_filter') == 'yes'

    # Tính dãy ngày
    start = datetime.strptime(start_date, '%Y-%m-%d').date()
    end = datetime.strptime(end_date, '%Y-%m-%d').date()
    days_in_range = [start + timedelta(days=i) for i in range((end - start).days + 1)]

    # Truy vấn người dùng
    query = User.query
    if department:
        query = query.filter(User.department == department)

    # Lọc theo loại hợp đồng nếu cần
    if selected_contract:
        if selected_contract.lower() == "hợp đồng":
            query = query.filter(or_(
                User.contract_type.ilike("hợp đồng%"),
                User.contract_type.ilike("%hợp đồng"),
                User.contract_type.ilike("%hợp đồng%")
            ))
        else:
            query = query.filter(User.contract_type.ilike(selected_contract))

    users = query.order_by(User.name).all()

    # Lấy lịch trực
    schedules = Schedule.query.join(Shift).filter(
        Schedule.user_id.in_([u.id for u in users]),
        Schedule.work_date.between(start, end)
    ).all()

    # Ánh xạ lịch trực và thống kê công
    schedule_map = {}
    summary = {user.id: {'kl': 0, 'tg': 0, '100': 0, 'bhxh': 0} for user in users}
    for s in schedules:
        key = (s.user_id, s.work_date)
        code = s.shift.code.upper() if s.shift and s.shift.code else 'X'
        schedule_map[key] = code

        shift_name = s.shift.name.lower()
        if shift_name in ['nghỉ', 'h']:
            summary[s.user_id]['kl'] += 1
        elif shift_name == 'bhxh':
            summary[s.user_id]['bhxh'] += 1
        elif shift_name == '100':
            summary[s.user_id]['100'] += 1
        else:
            summary[s.user_id]['tg'] += 1

    holidays = [
        date(2025, 1, 1),
        date(2025, 4, 30),
        date(2025, 5, 1),
        date(2025, 9, 2),
    ]

    departments = [d[0] for d in db.session.query(User.department).filter(User.department != None).distinct().all()]
    now = datetime.today()

    return render_template("bang_cham_cong.html",
                           users=users,
                           days_in_month=days_in_range,
                           schedule_map=schedule_map,
                           summary=summary,
                           month=start.month,
                           year=start.year,
                           departments=departments,
                           selected_department=department,
                           selected_contract=selected_contract,
                           print_filter=print_filter,
                           start_date=start_date,
                           end_date=end_date,
                           holidays=holidays,
                           now=now)

from flask import render_template, request, send_file
from datetime import datetime, timedelta
from models import User, Shift, Schedule
import openpyxl
from io import BytesIO

@app.route('/report-all')
def report_all():
    start_str = request.args.get('start')
    end_str = request.args.get('end')
    if start_str and end_str:
        start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()
        end_date = start_date + timedelta(days=6)

    schedules = Schedule.query.join(User).join(Shift) \
        .filter(Schedule.work_date.between(start_date, end_date)) \
        .order_by(Schedule.user_id, Schedule.work_date).all()

    print(f"[DEBUG] Found {len(schedules)} schedule entries")  # 🚧 debug

    grouped = {}
    for s in schedules:
        key = (s.user.department or 'Khác', s.user.position or '')
        day_key = s.work_date.strftime('%a %d/%m')
        grouped.setdefault(key, {})
        grouped[key].setdefault(day_key, "")
        grouped[key][day_key] += f"{s.user.name} ({s.shift.name})\n"

    date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

    print("[DEBUG] Grouped keys:", grouped.keys())

    return render_template('report_all.html',
                           grouped=grouped,
                           date_range=date_range,
                           start_date=start_date,
                           end_date=end_date)


@app.route('/export-report-all')
def export_report_all():
    start_str = request.args.get('start')
    end_str = request.args.get('end')
    if start_str and end_str:
        start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
    else:
        start_date = datetime.today().date()
        end_date = start_date + timedelta(days=6)

    schedules = Schedule.query.join(User).join(Shift) \
        .filter(Schedule.work_date.between(start_date, end_date)) \
        .order_by(Schedule.work_date).all()

    grouped = {}
    for s in schedules:
        dept = s.user.department or 'Khác'
        pos = s.user.position or ''
        key = (dept, pos)
        grouped.setdefault(key, {})
        day = s.work_date.strftime('%a %d/%m')
        grouped[key][day] = grouped[key].get(day, '') + f"{s.user.name} ({s.shift.name}); "

    date_range = [(start_date + timedelta(days=i)).strftime('%a %d/%m')
                  for i in range((end_date - start_date).days + 1)]

    wb = openpyxl.Workbook()
    ws = wb.active
    header = ['Khoa/Phòng', 'Chức danh'] + date_range
    ws.append(header)

    for (dept, pos), days in grouped.items():
        row = [dept, pos] + [days.get(d, '') for d in date_range]
        ws.append(row)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(stream, as_attachment=True, download_name='lich_truc_toan_vien.xlsx')

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)
import seed  # tự động gọi seed.py khi khởi chạy server
